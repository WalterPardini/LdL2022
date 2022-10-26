'''
5 - Un script Python que sea capaz de crear una planilla de cálculo a modo de reporte con
los datos de una campaña publicitaria finalizada dada. Este script, además debe enviar
por mail el reporte creado al cliente en cuestión.
'''

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference, PieChart,ProjectedPieChart
import string

informes = pd.read_excel("BDE.xlsx")
tabla_pivote = informes.pivot_table(index="MODALIDAD",columns='OPERADOR', values='BLOQUE')
tabla_pivote.to_excel('Informe Automatico - Danna Fox.xlsx', startrow=1,startcol=1,sheet_name='Octubre 2022')

#Cargar Libro Excel
lectura = load_workbook('Informe Automatico - Danna Fox.xlsx')
pestaña = lectura['Octubre 2022']
min_columna = lectura.active.min_column
max_columna = lectura.active.max_column
min_fila = lectura.active.min_row
max_fila = lectura.active.max_row


#Grafico de Barras
graficoDeBarras = BarChart()
data = Reference(pestaña, min_col=min_columna+1,max_col=max_columna, min_row=min_fila,max_row=max_fila)
categorias = Reference(pestaña, min_col=min_columna,max_col=min_columna, min_row=min_fila+1,max_row=max_fila)

graficoDeBarras.add_data(data, titles_from_data=True)
graficoDeBarras.set_categories(categorias)
graficoDeBarras.title = "Resumen"
graficoDeBarras.y_axis.title = 'Cantidad de Usuarios'
graficoDeBarras.x_axis.title = 'Modalidad'
graficoDeBarras.width = 30
graficoDeBarras.height = 15
pestaña.add_chart(graficoDeBarras,"B7")

'''
#Formulas de Excel
pestaña["B5"] = "Recuento"
i="B"
columnas = list(string.ascii_uppercase)
sumas = columnas[0:max_columna]
for i in sumas:
    if i!="U":
        pestaña[f'{i}5'] = f'=SUM({i}3:{i}4'
'''     

#Guarda los Cambios Realizados
lectura.save('Informe Automatico - Danna Fox.xlsx')